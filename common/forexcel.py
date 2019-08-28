import openpyxl
import os

class Case:

    def __init__(self): # 初始化,先传个空值,读取时候再赋值
        self.method = None
        self.url = None


class DoExcel:

        def __init__(self,file_name):
            try:
                self.workbook = openpyxl.load_workbook(filename=file_name)

            except FileNotFoundError as e:
                print("Excel不存在:{0}".format(e))
                raise e

        def read_cases(self,sheet_name):
            sheet = self.workbook[sheet_name]
            max_row = sheet.max_row#最大行
            test_case = []
            for i in range(2,max_row+1):
                excel_case = Case()
                excel_case.id = sheet.cell(row=i, column=1).value
                excel_case.title = sheet.cell(row=i, column=2).value
                excel_case.method = sheet.cell(row=i, column=3).value
                excel_case.url = sheet.cell(row=i, column=4).value
                excel_case.data = sheet.cell(row=i, column=5).value
                test_case.append(excel_case)  # 读取出来的数据添加到list
            return test_case


        def read_sheet(self):
            return self.workbook.sheetnames  #获取所有表单


if __name__ == '__main__':
    '''
    os.path.join路径组合返回
    os.path.split()：按照路径将文件名和路径分割开
    os.path.realpath(__file__)获取文件所在目录
    '''
    excel_path = os.path.join(os.path.split(os.path.split(os.path.realpath(__file__))[0])[0],
                              r'data\test_case_web.xlsx')
    print(excel_path)
    t = DoExcel(excel_path).read_cases('web')
    print(t)
    for case_run in t:  # 遍历出组装好的测试用例
        print('\ntest_case信息:{}'.format(case_run.__dict__))

